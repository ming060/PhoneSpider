from openpyxl import Workbook
import json

if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    ws.title = "All"
    ws.cell(1, 1, 'Brand')
    ws.cell(1, 2, 'Model')
    ws.cell(1, 3, 'Alias')

    with open('../PhoneSpider/phones.json') as f:
        devices = json.load(f)
        for index, device in enumerate(devices):
            ws.cell(row = index+2, column = 1, value = device['brand'][0])
            ws.cell(row = index+2, column = 2, value = device['model'][0])
            aliases = device['alias']
            for alias_index, alias in enumerate(aliases):
                ws.cell(row = index+2, column = alias_index + 3, value = alias)

    wb.save('PhoneSpecs.xlsx')
